from tkinter import *
from tkinter.ttk import *
from openpyxl import workbook, load_workbook
import random
root = Tk()
root.title("Randomization lottery")
root.geometry("300x300")
root.resizable(0, 0)   # people can't expand

def randomname(event):
    wb = load_workbook('hello.xlsx')
    ws = wb.active
    rangeline = ws['A2': 'A19']  # we are creating a list here
    # our list is names as rangeline above

    name = []  # creating an empty variable ready to take LIST
    for items in rangeline:
        for subitems in items:
            name.append(subitems.value)
    computer_action = random.choice(name)
    print("The computer has chosen: " + computer_action)

    updateDisplay(computer_action)

def updateDisplay(abc):
    displayVariable.set(abc)

button_1 = Button(root, text = "Random Lottery Name")
button_1.bind("<Button-1>", randomname)
button_1.pack()
displayVariable = StringVar()
displayLabel = Label(root, textvariable = displayVariable)
displayLabel.pack()

mainloop()